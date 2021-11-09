'''
	Developed by Ricardo Dagnoni Huelsmann
	  Santa Catarina Federal University
		    Chemistry Department
				    2021
'''


# First of all, some libraries must be imported.
import threading
import serial
import serial.tools.list_ports
import time
from datetime import datetime
from tkinter import *
from openpyxl import Workbook
import sys

state = True

''' 
Starting the creation of some functions, a smart one is able to auto-find and auto-connect
the software with the Arduino Uno, through serial communication port at the same 9600 bps 
described in the Arduino sketch. This function uses the 'serial.tools.list_ports' library
to find the ports in use by the computer, find the one used by Arduino, splits its name and creates 
the string "COM3" (or COM4, COM6 for example) and them starts the serial communication.
It is important to explain that this code works, because the presence of 'Arduino' word in the name 
of the serial port. It will not work in the case of Arduino Nano, because the port name has no 'Arduino' on it 
'''


def auto_connect_arduino():
    global ser                                            # A global variable 'ser' (related to serial) is defined, being able to be used in others functions
    computer_ports = serial.tools.list_ports.comports()   # The ports in use by the computed are find and stored in the variable 'computer_ports'
    for i in range(len(computer_ports)):                  # A 'for' loop is started. It will be executed 'i' times, which is the lenght of the computer_ports. In other words, it will check every COM por used.
        port = str(computer_ports[i])					  # A variable 'port' is created, containing a string, which is the COM port.
        if 'Arduino' in port:							  # The 'port' is checked. If there are 'Arduino' word on it, the code below is executed.
            split_port = port.split(' ')				  # The 'port' string is splitted on every blank character. 
            communication_port = (split_port[0])          # As the name of the port get is in the form 'COM7 - Arduino Uno (COM7)', the first part of the split method gets exactely the name of the COM port.
            ser = serial.Serial(communication_port, 9600) # The serial communication is then created, in the name of 'ser'

auto_connect_arduino()									  # This line of the code executes the previous declared function.



'''The next functions make use of the datetime library to find the current date and time, and return specific information.'''


def get_file_name():            			              # This function return a file name (string), based on the current time and date
	full_date_time = datetime.now()						  # The function 'now()' in the datetime, gets a full date and time information
	year = str(full_date_time.year)						  # This line gets specifically the current year, converting it to a string
	month = str(full_date_time.month)					  # This line gets specifically the current month, converting it to a string
	day = str(full_date_time.day)			  	          # This line gets specifically the current day, converting it to a string
	hour = str(full_date_time.hour)					      # This line gets specifically the current hour, converting it to a string
	minute = str(full_date_time.minute)                   # This line gets specifically the current minute, converting it to a string
	file_name = (month + "_" + day + "_" + year + '_' + hour + '_' + minute + ".xlsx")       # The variable file_name is created, appending datetime information with underscores, and defining the file type at the end (.xlsx)
	return file_name                                      # In the end, returns the file name obtained

name_initial_time = get_file_name()						  # In order to create the file name as soon as possible, this code execute the previous function and stores the file name in the proper variable.

def get_date():                                           # This function return a full date information (string), based on the current time and date
	full_date_time = datetime.now()                       # This lines below work in the same way as the previous one
	year = str(full_date_time.year)
	month = str(full_date_time.month)
	day = str(full_date_time.day)
	full_date = (month + "/" + day + "/" + year)
	return full_date

def get_time():											  # Similarly to the previous function, this one return a full time information (string)
	full_date_time = datetime.now()						  # This lines below work in the same way as the previous one
	hour = str(full_date_time.hour)
	minute = full_date_time.minute
	if (minute<10):										  # In this specific line, a correction is made if the minute information is below 10
		minute = "0" + str(minute)						  # A character '0' have to be inserted to achieve a two digit string, like "04" or "06"
	else:											      # If the current minute is 10 or higher, it only have to be converted to string
		minute = str(minute)
	second = full_date_time.second
	if (second<10):										  # The same process have to be made with the information related to the seconds
		second = "0" + str(second)
	else:
		second = str(second)
	full_time = hour + ":" + minute + ':' + second
	return full_time


def get_temperature():									  # This function read the data from Arduino (DHT sensor) and return the current temperature
	received_byte = ser.readline()					      # Initially, the information is read from the serial communication and stored in the 'received_byte' variable
	received_string = str(received_byte)				  # The received byte is converted to a string and stored in the 'receved_string' variable
	temperature = received_string[18:23]				  # The temperature information is obtained from within the received string (corresponding to a fraction of it) and stored in the proper variable
	return temperature 									  # Return the information previously obtained

def get_humi():											  # In the same way as the get_temperature function, this one read the data from Arduino and return the current humidity
	received_byte = ser.readline()						  # Read the serial information and store it
	received_string = str(received_byte)				  # Convert to a string
	humidity = received_string[7:12]					  # In this case, as expected, the inforation about humidity is accessed in a diffent part of the string
	return humidity 	                                  # Return the information previously obtained



'''The next functions make use of the tkinter and openpyxl libraries.'''

def start():						# This function is accessed when a button is pressed in the main software window (it will be discussed below in the 'tkinter' section)
	global label_1 					# Define 'label' as a global variable, since it will be used in many ways
	button_1.pack_forget()          # Erases button 1 in the screen
	label_1.pack_forget()			# Erases label 1 in the screen (the label 1 is defined and packed below)
	label_1 = Label(window, text="\n Hello, welcome to the Temperature and Humidity \n data collection software! \n \n Press the end button to finish the measurement \n \n")   # Redefine label 1, in the main window, changing its text
	label_1.pack()  				# Puts the label 1 on the screen
	button_2.pack()					# Puts the button 2 on the screen
	number = 3						# Creates a variable called 'number', and store the number 3 in it. This number is related to the starting line foor the Excel data writing
	try: 							# Uses a try/except function to execute an infite loop "while True". In the end of the measurement, the infinite loop will not run properly because the serial communication will be closed, and the except line will run.
		while True:			   	 	# Creates a infinite loop to collect information periodically from the Arduino, with no timeout.
			sheet1.cell(row=number, column=2).value = get_date()
			sheet1.cell(row=number, column=3).value = get_time()
			sheet1.cell(row=number, column=4).value = get_temperature()
			sheet1.cell(row=number, column=5).value = get_humi()
			real_number = number-2
			print('Collected data: ' + str(real_number))
			workbook.save('safety_version.xlsx')
			number += 1
			for i in range (30):
				if state == False:
						window.quit()
						sys.exit()
				time.sleep(2)
	except:
		window.quit()
				

def end():
	global state 
	ser.close()
	workbook.save(name_initial_time)
	button_2.pack_forget()
	label_1.pack_forget()
	label_2 = Label(window, text="\n \n Thank you! \n\n The program is finishing \n\n You can close the window or wait for the program end")
	label_2.pack()
	time.sleep(2)
	state = False


# The following part of the code concerns the creation and formatting of the Excel spreadsheet
workbook = Workbook()                                     # Using 'openpyxl' library, an Excel workbook is created.
sheet1 = workbook.worksheets[0]						      # Then, one sheet is created in the workbook.
headers = ['Date', 'Time', 'Temperature', 'Humidity']	  # This line creates an array containing the headers of the table.
for i in range (2,6):									  # A for loop is used to insert each header name in a proper cell in the Excel worksheet. The range starts in 2, because is the column position for the first header, and goes to six, encompassing all the headers.
	sheet1.cell(row=2, column=i).value = headers[i-2]	  # Each cell is filled with a header name from the array above defined. All of them will be at the second row, but the colunm is changed for each case. The 'i-2' inside the brackets is need to acess the right header position in the array (from 0 to 3)
	

#Tkinter informations
window = Tk()
window.title("Temperature and humidity informations")
label_1 = Label(window, text="\n Hello, welcome to the Temperature and Humidity \n data collection software! \n \n Press the start button to initiate the measurement \n \n")
label_1.pack()
window.geometry('450x180')
button_1 = Button(window, text="Start measurement", command=threading.Thread(target=start).start)
button_1.pack()
button_2 = Button(window, text='End process', command=threading.Thread(target=end).start)



window.mainloop()